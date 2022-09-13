import sys
import os
import openpyxl
import datetime

def memo(sys):
    fname = "fmemo.txt"
    option = sys.argv[1]
    if option == "-a":
        if len(sys.argv) > 2:
            memotxt = sys.argv[2]
            with open(fname,"a") as f:
                f.write(memotxt+"\n")
        else:
            print("no add memo")
    elif option == "-v":
        try:
            file = open(fname,"r")
            memotxt = file.read()
            file.close()
            print(memotxt)
        except:
            None

def search(sys):
    if len(sys.argv) == 2:
        dirname = sys.argv[1]
        fileext = ""
    if len(sys.argv) == 3:
        dirname = sys.argv[1]
        fileext = sys.argv[2]
    
    if dirname != "":
        filenames = os.listdir(dirname)
        for filename in filenames:
            full_filename = os.path.join(dirname,filename)
            ext = os.path.splitext(full_filename)[-1]
            if fileext != "":
                if(ext == fileext):
                    print(full_filename)
            else:
                print(full_filename)

def makexlsfile():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "first data"
    ws['A2'] = datetime.datetime.now()
    ws.append([1,2,3])
    wb.save("sample.xlsx")

def readxlsfile():
    wb = openpyxl.load_workbook("sample.xlsx")
    ws = wb.active
    xlsdata = []
    for r in range(1,ws.max_row+1):
        celdata = []
        for c in range(1,ws.max_column+1):
            print("r:{}, c:{}, value:{}".format(r,c,ws.cell(r,c).value))
            celdata.append(ws.cell(r,c).value)
        xlsdata.append(celdata)
    print(xlsdata)

if len(sys.argv) > 1:
    if sys.argv[1][0] == '-':
        memo(sys)
    else:
        search(sys)

readxlsfile()
